import os
import logging
import sqlite3
import threading
import json
import io
from datetime import datetime, timedelta, timezone
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, ContextTypes, CommandHandler, CallbackQueryHandler
from telegram.request import HTTPXRequest
from flask import Flask, render_template_string, send_file, jsonify, request as flask_request
import openpyxl
from waitress import serve

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
CHANNEL_USERNAME = os.getenv("CHANNEL_USERNAME")
MANAGER_USERNAME = os.getenv("MANAGER_USERNAME")

try:
    ADMIN_ID = int(os.getenv("ADMIN_ID", "0"))
except ValueError:
    ADMIN_ID = 0
    print("WARNING: ADMIN_ID in .env is not a valid number.")

DB_NAME = "referrals.db"

# ─────────────────────────────────────────────
#  DATABASE
# ─────────────────────────────────────────────

def init_db():
    with sqlite3.connect(DB_NAME) as conn:
        c = conn.cursor()

        # Core tables
        c.execute('''CREATE TABLE IF NOT EXISTS referrals
                     (user_id INTEGER PRIMARY KEY,
                      referrer_id INTEGER,
                      username TEXT,
                      first_name TEXT,
                      join_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

        # Migration: add new columns to existing databases that predate this version
        existing_cols = {row[1] for row in c.execute("PRAGMA table_info(referrals)")}
        if "username" not in existing_cols:
            c.execute("ALTER TABLE referrals ADD COLUMN username TEXT")
            print("DB Migration: added 'username' column to referrals")
        if "first_name" not in existing_cols:
            c.execute("ALTER TABLE referrals ADD COLUMN first_name TEXT")
            print("DB Migration: added 'first_name' column to referrals")

        c.execute('''CREATE TABLE IF NOT EXISTS purchases
                     (id INTEGER PRIMARY KEY AUTOINCREMENT,
                      user_id INTEGER,
                      product_name TEXT,
                      amount REAL,
                      purchase_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                      FOREIGN KEY(user_id) REFERENCES referrals(user_id))''')

        # New: link click tracking
        c.execute('''CREATE TABLE IF NOT EXISTS link_clicks
                     (id INTEGER PRIMARY KEY AUTOINCREMENT,
                      referrer_id INTEGER,
                      clicked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

        # New: content posts tracking (who shared content)
        c.execute('''CREATE TABLE IF NOT EXISTS content_posts
                     (id INTEGER PRIMARY KEY AUTOINCREMENT,
                      user_id INTEGER,
                      post_type TEXT,
                      posted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')

        conn.commit()


def get_referrer(user_id):
    try:
        with sqlite3.connect(DB_NAME) as conn:
            c = conn.cursor()
            c.execute("SELECT referrer_id FROM referrals WHERE user_id=?", (user_id,))
            result = c.fetchone()
            return result[0] if result else None
    except Exception as e:
        print(f"DB Error: {e}")
        return None


def record_referral(user_id, referrer_id, username=None, first_name=None):
    try:
        with sqlite3.connect(DB_NAME) as conn:
            c = conn.cursor()
            c.execute(
                "INSERT INTO referrals (user_id, referrer_id, username, first_name) VALUES (?, ?, ?, ?)",
                (user_id, referrer_id, username, first_name)
            )
            # Count this as a click on the referrer's link
            c.execute("INSERT INTO link_clicks (referrer_id) VALUES (?)", (referrer_id,))
            conn.commit()
            return True
    except sqlite3.IntegrityError:
        return False
    except Exception as e:
        print(f"DB Error: {e}")
        return False


def record_new_user(user_id, username=None, first_name=None):
    """Record organic user (no referrer)."""
    try:
        with sqlite3.connect(DB_NAME) as conn:
            c = conn.cursor()
            c.execute(
                "INSERT OR IGNORE INTO referrals (user_id, referrer_id, username, first_name) VALUES (?, NULL, ?, ?)",
                (user_id, username, first_name)
            )
            conn.commit()
    except Exception as e:
        print(f"DB Error: {e}")


def log_purchase(user_id, product, amount):
    with sqlite3.connect(DB_NAME) as conn:
        c = conn.cursor()
        c.execute(
            "INSERT INTO purchases (user_id, product_name, amount) VALUES (?, ?, ?)",
            (user_id, product, amount)
        )
        conn.commit()


def log_content_post(user_id, post_type="link_share"):
    with sqlite3.connect(DB_NAME) as conn:
        c = conn.cursor()
        c.execute(
            "INSERT INTO content_posts (user_id, post_type) VALUES (?, ?)",
            (user_id, post_type)
        )
        conn.commit()


# ─────────────────────────────────────────────
#  ANALYTICS QUERIES
# ─────────────────────────────────────────────

def get_dashboard_stats():
    with sqlite3.connect(DB_NAME) as conn:
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        now = datetime.now(timezone.utc).replace(tzinfo=None)  # naive UTC, matches DB strings
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        week_start = now - timedelta(days=now.weekday())

        # Total clicks
        c.execute("SELECT COUNT(*) as cnt FROM link_clicks")
        total_clicks = c.fetchone()["cnt"]

        # Total buyers
        c.execute("SELECT COUNT(DISTINCT user_id) as cnt FROM purchases")
        total_buyers = c.fetchone()["cnt"]

        # Conversion rate
        conversion = round((total_buyers / total_clicks * 100), 1) if total_clicks > 0 else 0

        # Average purchase amount
        c.execute("SELECT AVG(amount) as avg FROM purchases")
        avg_purchase = c.fetchone()["avg"] or 0
        avg_purchase = round(avg_purchase, 2)

        # Total revenue
        c.execute("SELECT SUM(amount) as total FROM purchases")
        total_revenue = c.fetchone()["total"] or 0
        total_revenue = round(total_revenue, 2)

        # Purchase count
        c.execute("SELECT COUNT(*) as cnt FROM purchases")
        purchase_count = c.fetchone()["cnt"]

        # New referrals this month
        c.execute("SELECT COUNT(*) as cnt FROM referrals WHERE join_date >= ?", (month_start.isoformat(),))
        monthly_referrals = c.fetchone()["cnt"]

        # Total referrals
        c.execute("SELECT COUNT(*) as cnt FROM referrals")
        total_referrals = c.fetchone()["cnt"]

        # Revenue this month
        c.execute("SELECT SUM(amount) as s FROM purchases WHERE purchase_date >= ?", (month_start.isoformat(),))
        monthly_revenue = c.fetchone()["s"] or 0
        monthly_revenue = round(monthly_revenue, 2)

        # Top referrers (all time) - who brought most clients
        c.execute('''
            SELECT r.referrer_id,
                   COUNT(r.user_id) as invite_count,
                   COALESCE(ref_info.username, CAST(r.referrer_id AS TEXT)) as display_name,
                   COALESCE(SUM(p.amount), 0) as total_earned
            FROM referrals r
            LEFT JOIN referrals ref_info ON ref_info.user_id = r.referrer_id
            LEFT JOIN purchases p ON p.user_id = r.user_id
            WHERE r.referrer_id IS NOT NULL
            GROUP BY r.referrer_id
            ORDER BY invite_count DESC
            LIMIT 10
        ''')
        top_referrers = [dict(row) for row in c.fetchall()]

        # Average earnings per referral (for referrers who have earnings)
        c.execute('''
            SELECT AVG(earnings) as avg_earn FROM (
                SELECT r.referrer_id, SUM(p.amount) as earnings
                FROM referrals r
                JOIN purchases p ON p.user_id = r.user_id
                WHERE r.referrer_id IS NOT NULL
                GROUP BY r.referrer_id
            )
        ''')
        avg_earning_per_ref = c.fetchone()["avg_earn"] or 0
        avg_earning_per_ref = round(avg_earning_per_ref, 2)

        # Monthly referral breakdown (last 6 months)
        monthly_data = []
        for i in range(5, -1, -1):
            month_dt = (now.replace(day=1) - timedelta(days=i*28)).replace(day=1)
            next_month = (month_dt.replace(day=28) + timedelta(days=4)).replace(day=1)
            c.execute("SELECT COUNT(*) as cnt FROM referrals WHERE join_date >= ? AND join_date < ?",
                      (month_dt.isoformat(), next_month.isoformat()))
            cnt = c.fetchone()["cnt"]
            c.execute("SELECT COALESCE(SUM(amount),0) as s FROM purchases WHERE purchase_date >= ? AND purchase_date < ?",
                      (month_dt.isoformat(), next_month.isoformat()))
            rev = c.fetchone()["s"]
            monthly_data.append({
                "month": month_dt.strftime("%b %Y"),
                "referrals": cnt,
                "revenue": round(rev, 2)
            })

        # Top flower of the week (most referrals this week)
        c.execute('''
            SELECT r.referrer_id,
                   COUNT(r.user_id) as invite_count,
                   COALESCE(ref_info.username, CAST(r.referrer_id AS TEXT)) as display_name
            FROM referrals r
            LEFT JOIN referrals ref_info ON ref_info.user_id = r.referrer_id
            WHERE r.referrer_id IS NOT NULL AND r.join_date >= ?
            GROUP BY r.referrer_id
            ORDER BY invite_count DESC
            LIMIT 1
        ''', (week_start.isoformat(),))
        flower_week = c.fetchone()
        flower_week = dict(flower_week) if flower_week else None

        # Top flower of the month
        c.execute('''
            SELECT r.referrer_id,
                   COUNT(r.user_id) as invite_count,
                   COALESCE(ref_info.username, CAST(r.referrer_id AS TEXT)) as display_name
            FROM referrals r
            LEFT JOIN referrals ref_info ON ref_info.user_id = r.referrer_id
            WHERE r.referrer_id IS NOT NULL AND r.join_date >= ?
            GROUP BY r.referrer_id
            ORDER BY invite_count DESC
            LIMIT 1
        ''', (month_start.isoformat(),))
        flower_month = c.fetchone()
        flower_month = dict(flower_month) if flower_month else None

        # Most active content posters
        c.execute('''
            SELECT cp.user_id,
                   COUNT(*) as post_count,
                   COALESCE(r.username, CAST(cp.user_id AS TEXT)) as display_name
            FROM content_posts cp
            LEFT JOIN referrals r ON r.user_id = cp.user_id
            GROUP BY cp.user_id
            ORDER BY post_count DESC
            LIMIT 5
        ''')
        top_posters = [dict(row) for row in c.fetchall()]

        # Monthly payout per referrer (this month)
        c.execute('''
            SELECT r.referrer_id,
                   COALESCE(ref_info.username, CAST(r.referrer_id AS TEXT)) as display_name,
                   SUM(p.amount) as monthly_total,
                   COUNT(p.id) as sale_count
            FROM referrals r
            JOIN purchases p ON p.user_id = r.user_id
            LEFT JOIN referrals ref_info ON ref_info.user_id = r.referrer_id
            WHERE r.referrer_id IS NOT NULL AND p.purchase_date >= ?
            GROUP BY r.referrer_id
            ORDER BY monthly_total DESC
            LIMIT 10
        ''', (month_start.isoformat(),))
        monthly_payouts = [dict(row) for row in c.fetchall()]

        # Recent sales
        c.execute('''
            SELECT p.user_id, p.product_name, p.amount, r.referrer_id,
                   p.purchase_date,
                   COALESCE(ref_info.username, CAST(r.referrer_id AS TEXT)) as referrer_name
            FROM purchases p
            LEFT JOIN referrals r ON p.user_id = r.user_id
            LEFT JOIN referrals ref_info ON ref_info.user_id = r.referrer_id
            ORDER BY p.purchase_date DESC
            LIMIT 10
        ''')
        recent_sales = [dict(row) for row in c.fetchall()]

        return {
            "total_clicks": total_clicks,
            "total_buyers": total_buyers,
            "conversion": conversion,
            "avg_purchase": avg_purchase,
            "total_revenue": total_revenue,
            "purchase_count": purchase_count,
            "monthly_referrals": monthly_referrals,
            "total_referrals": total_referrals,
            "monthly_revenue": monthly_revenue,
            "top_referrers": top_referrers,
            "avg_earning_per_ref": avg_earning_per_ref,
            "monthly_data": monthly_data,
            "flower_week": flower_week,
            "flower_month": flower_month,
            "top_posters": top_posters,
            "monthly_payouts": monthly_payouts,
            "recent_sales": recent_sales,
        }


# ─────────────────────────────────────────────
#  TELEGRAM BOT
# ─────────────────────────────────────────────

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.error(msg="Exception while handling an update:", exc_info=context.error)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    args = context.args

    if args and args[0].isdigit():
        potential_referrer = int(args[0])
        if potential_referrer != user.id:
            is_new = record_referral(
                user.id, potential_referrer,
                username=user.username,
                first_name=user.first_name
            )
            if is_new:
                try:
                    await context.bot.send_message(
                        chat_id=potential_referrer,
                        text=f"🎉 New Referral! {user.first_name} just joined via your link."
                    )
                except Exception:
                    pass
    else:
        record_new_user(user.id, username=user.username, first_name=user.first_name)

    keyboard = [
        [InlineKeyboardButton("📢 Join Channel", url=f"https://t.me/{CHANNEL_USERNAME.replace('@', '')}")],
        [InlineKeyboardButton("🔗 Get My Referral Link", callback_data="get_link")],
        [InlineKeyboardButton("🛒 Buy Premium Plan", callback_data="buy_plan")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    welcome_text = (
        f"Hello {user.first_name}! Welcome to the official bot.\n\n"
        "We track referrals and rewards here. "
        "Join our channel for updates or generate your own link to invite friends!"
    )

    await update.message.reply_text(welcome_text, reply_markup=reply_markup)


async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user = query.from_user
    await query.answer()

    if query.data == "get_link":
        bot_username = context.bot.username
        ref_link = f"https://t.me/{bot_username}?start={user.id}"
        # Log that they shared their link (content activity)
        log_content_post(user.id, "link_share")

        msg = (
            f"Here is your unique referral link:\n\n<code>{ref_link}</code>\n\n"
            "Share this on TikTok or Instagram. When people join via this link, "
            "we will know you sent them!"
        )
        await query.edit_message_text(text=msg, parse_mode='HTML')

    elif query.data == "buy_plan":
        msg = (
            f"🛍️ <b>How to Purchase:</b>\n\n"
            f"Please contact our manager {MANAGER_USERNAME} to complete your payment.\n\n"
            f"⚠️ <b>Important:</b> Please send them your User ID so they can process your order.\n\n"
            f"Your User ID is: <code>{user.id}</code>"
        )
        await query.edit_message_text(text=msg, parse_mode='HTML')


async def approve_sale(update: Update, context: ContextTypes.DEFAULT_TYPE):
    admin_user = update.effective_user

    if admin_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ You are not authorized to use this command.")
        return

    try:
        target_user_id = int(context.args[0])
        amount = float(context.args[1])
        product = " ".join(context.args[2:]) if len(context.args) > 2 else "Premium Plan"
    except (IndexError, ValueError):
        await update.message.reply_text(
            "❌ Usage: /approve <user_id> <amount> [product_name]\n"
            "Example: <code>/approve 123456789 49.99 Premium Plan</code>",
            parse_mode='HTML'
        )
        return

    log_purchase(target_user_id, product, amount)
    referrer_id = get_referrer(target_user_id)

    response_text = f"✅ Sale approved for User ID {target_user_id} (${amount} — {product}).\n"

    if referrer_id:
        response_text += f"Attribution: referred by {referrer_id}. Notifying them now."
        try:
            await context.bot.send_message(
                chat_id=referrer_id,
                text=f"💰 Cha-ching! Your referral (User ID {target_user_id}) just purchased '{product}' for ${amount}. You earned credit!"
            )
        except Exception:
            pass
    else:
        response_text += "Attribution: Organic (no referrer)."

    try:
        await context.bot.send_message(
            chat_id=target_user_id,
            text="🎉 Your purchase has been approved! Thank you for your business."
        )
    except Exception:
        pass

    await update.message.reply_text(response_text)


async def log_post_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin command: /logpost <user_id> <type> — manually log a content post."""
    admin_user = update.effective_user
    if admin_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ Not authorized.")
        return
    try:
        uid = int(context.args[0])
        post_type = context.args[1] if len(context.args) > 1 else "manual"
        log_content_post(uid, post_type)
        await update.message.reply_text(f"✅ Content post logged for user {uid} ({post_type}).")
    except (IndexError, ValueError):
        await update.message.reply_text("❌ Usage: /logpost <user_id> <post_type>")


# ─────────────────────────────────────────────
#  FLASK DASHBOARD
# ─────────────────────────────────────────────

app = Flask(__name__)

DASHBOARD_HTML = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ReferralBot · Analytics</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@300;400;500&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  :root {
    --bg:      #0b0c10;
    --surface: #13151c;
    --card:    #191c27;
    --border:  #252837;
    --accent:  #7c6af7;
    --accent2: #f76a6a;
    --accent3: #43e8a4;
    --accent4: #f5c842;
    --text:    #e8eaf2;
    --muted:   #6e7191;
    --font-head: 'Syne', sans-serif;
    --font-mono: 'DM Mono', monospace;
  }

  html { scroll-behavior: smooth; }

  body {
    background: var(--bg);
    color: var(--text);
    font-family: var(--font-mono);
    min-height: 100vh;
    overflow-x: hidden;
  }

  /* Noise overlay */
  body::before {
    content: '';
    position: fixed; inset: 0;
    background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)' opacity='0.04'/%3E%3C/svg%3E");
    pointer-events: none;
    z-index: 999;
    opacity: 0.4;
  }

  /* ── Header ── */
  header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 1.25rem 2rem;
    background: var(--surface);
    border-bottom: 1px solid var(--border);
    position: sticky; top: 0;
    z-index: 100;
    backdrop-filter: blur(12px);
  }

  .logo {
    font-family: var(--font-head);
    font-size: 1.15rem;
    font-weight: 800;
    letter-spacing: -0.02em;
    display: flex; align-items: center; gap: .6rem;
  }

  .logo-dot {
    width: 8px; height: 8px;
    background: var(--accent3);
    border-radius: 50%;
    display: inline-block;
    animation: pulse 2s infinite;
  }

  @keyframes pulse {
    0%, 100% { opacity: 1; transform: scale(1); }
    50% { opacity: .5; transform: scale(1.4); }
  }

  .header-right { display: flex; align-items: center; gap: 1rem; }

  .badge-live {
    background: #0e2e1e;
    color: var(--accent3);
    font-size: .7rem;
    font-weight: 500;
    padding: .25rem .6rem;
    border-radius: 99px;
    border: 1px solid #1a5038;
    letter-spacing: .08em;
    text-transform: uppercase;
  }

  .btn-export {
    background: var(--accent);
    color: #fff;
    border: none;
    padding: .5rem 1.1rem;
    border-radius: 8px;
    font-family: var(--font-mono);
    font-size: .8rem;
    font-weight: 500;
    cursor: pointer;
    text-decoration: none;
    transition: opacity .2s, transform .1s;
    display: flex; align-items: center; gap: .4rem;
  }
  .btn-export:hover { opacity: .85; transform: translateY(-1px); }

  /* ── Layout ── */
  main { padding: 2rem; max-width: 1440px; margin: 0 auto; }

  .section-title {
    font-family: var(--font-head);
    font-size: .7rem;
    font-weight: 600;
    letter-spacing: .15em;
    text-transform: uppercase;
    color: var(--muted);
    margin-bottom: 1rem;
    margin-top: 2.5rem;
    display: flex; align-items: center; gap: .5rem;
  }
  .section-title::after {
    content: '';
    flex: 1;
    height: 1px;
    background: var(--border);
  }

  /* ── Stat Grid ── */
  .stat-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
    gap: 1rem;
  }

  .stat-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 1.4rem 1.5rem;
    position: relative;
    overflow: hidden;
    transition: border-color .2s, transform .2s;
  }
  .stat-card:hover { border-color: var(--accent); transform: translateY(-2px); }

  .stat-card::before {
    content: '';
    position: absolute;
    top: -40px; right: -40px;
    width: 100px; height: 100px;
    border-radius: 50%;
    opacity: .07;
  }

  .stat-card[data-accent="purple"]::before { background: var(--accent); }
  .stat-card[data-accent="red"]::before    { background: var(--accent2); }
  .stat-card[data-accent="green"]::before  { background: var(--accent3); }
  .stat-card[data-accent="yellow"]::before { background: var(--accent4); }

  .stat-label {
    font-size: .72rem;
    color: var(--muted);
    letter-spacing: .06em;
    text-transform: uppercase;
    margin-bottom: .6rem;
  }

  .stat-value {
    font-family: var(--font-head);
    font-size: 2rem;
    font-weight: 800;
    line-height: 1;
    color: var(--text);
  }

  .stat-sub {
    font-size: .72rem;
    color: var(--muted);
    margin-top: .4rem;
  }

  .stat-accent-purple { color: var(--accent); }
  .stat-accent-green  { color: var(--accent3); }
  .stat-accent-red    { color: var(--accent2); }
  .stat-accent-yellow { color: var(--accent4); }

  /* ── Flower cards ── */
  .flower-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
    gap: 1rem;
  }

  .flower-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 1.5rem;
    display: flex;
    align-items: center;
    gap: 1.2rem;
  }

  .flower-emoji {
    font-size: 2.5rem;
    line-height: 1;
  }

  .flower-info { flex: 1; min-width: 0; }
  .flower-period {
    font-size: .68rem;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: .1em;
  }
  .flower-name {
    font-family: var(--font-head);
    font-size: 1.2rem;
    font-weight: 700;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    margin-top: .2rem;
  }
  .flower-count {
    font-size: .78rem;
    color: var(--accent3);
    margin-top: .3rem;
  }

  /* ── Charts ── */
  .charts-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 1rem;
  }
  @media (max-width: 900px) { .charts-grid { grid-template-columns: 1fr; } }

  .chart-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 1.5rem;
  }

  .chart-title {
    font-family: var(--font-head);
    font-size: .85rem;
    font-weight: 700;
    color: var(--text);
    margin-bottom: 1.2rem;
  }

  /* ── Tables ── */
  .tables-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(420px, 1fr));
    gap: 1rem;
  }
  @media (max-width: 900px) { .tables-grid { grid-template-columns: 1fr; } }

  .table-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 14px;
    overflow: hidden;
  }

  .table-head {
    padding: 1.1rem 1.4rem;
    border-bottom: 1px solid var(--border);
    font-family: var(--font-head);
    font-size: .85rem;
    font-weight: 700;
    display: flex; align-items: center; justify-content: space-between;
  }

  table { width: 100%; border-collapse: collapse; }
  thead th {
    padding: .75rem 1.2rem;
    text-align: left;
    font-size: .68rem;
    font-weight: 500;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: .08em;
    border-bottom: 1px solid var(--border);
  }
  tbody tr { border-bottom: 1px solid var(--border); transition: background .15s; }
  tbody tr:last-child { border-bottom: none; }
  tbody tr:hover { background: rgba(124, 106, 247, .04); }
  tbody td { padding: .85rem 1.2rem; font-size: .82rem; }

  .pill {
    display: inline-block;
    padding: .2rem .55rem;
    border-radius: 99px;
    font-size: .7rem;
    font-weight: 500;
  }
  .pill-accent  { background: rgba(124,106,247,.18); color: var(--accent); }
  .pill-green   { background: rgba(67,232,164,.15);  color: var(--accent3); }
  .pill-yellow  { background: rgba(245,200,66,.15);  color: var(--accent4); }
  .pill-muted   { background: var(--border);          color: var(--muted); }

  .rank-num {
    font-family: var(--font-head);
    font-size: .8rem;
    font-weight: 700;
    color: var(--muted);
    width: 24px;
    display: inline-block;
    text-align: right;
  }
  .rank-num.top { color: var(--accent4); }

  .empty-row td {
    text-align: center;
    color: var(--muted);
    padding: 2.5rem;
    font-size: .82rem;
  }

  /* ── Footer ── */
  footer {
    text-align: center;
    padding: 2rem;
    color: var(--muted);
    font-size: .72rem;
    border-top: 1px solid var(--border);
    margin-top: 3rem;
  }

  /* ── Animations ── */
  @keyframes fadeUp {
    from { opacity: 0; transform: translateY(20px); }
    to   { opacity: 1; transform: translateY(0); }
  }
  .stat-card, .chart-card, .table-card, .flower-card {
    animation: fadeUp .4s ease both;
  }
  .stat-card:nth-child(1)  { animation-delay: .05s }
  .stat-card:nth-child(2)  { animation-delay: .10s }
  .stat-card:nth-child(3)  { animation-delay: .15s }
  .stat-card:nth-child(4)  { animation-delay: .20s }
  .stat-card:nth-child(5)  { animation-delay: .25s }
  .stat-card:nth-child(6)  { animation-delay: .30s }
  .stat-card:nth-child(7)  { animation-delay: .35s }
  .stat-card:nth-child(8)  { animation-delay: .40s }
</style>
</head>
<body>

<header>
  <div class="logo">
    <span class="logo-dot"></span>
    ReferralBot
  </div>
  <div class="header-right">
    <span class="badge-live">● Live</span>
    <a href="/export" class="btn-export">
      <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
      Export XLSX
    </a>
  </div>
</header>

<main>

  <!-- ── KPI Stats ── -->
  <div class="section-title">Key Metrics</div>
  <div class="stat-grid">
    <div class="stat-card" data-accent="purple">
      <div class="stat-label">Total Link Clicks</div>
      <div class="stat-value stat-accent-purple">{{ stats.total_clicks }}</div>
      <div class="stat-sub">All time</div>
    </div>
    <div class="stat-card" data-accent="green">
      <div class="stat-label">Conversion Rate</div>
      <div class="stat-value stat-accent-green">{{ stats.conversion }}%</div>
      <div class="stat-sub">Clicks → Purchases</div>
    </div>
    <div class="stat-card" data-accent="yellow">
      <div class="stat-label">Avg Purchase</div>
      <div class="stat-value stat-accent-yellow">${{ stats.avg_purchase }}</div>
      <div class="stat-sub">Per transaction</div>
    </div>
    <div class="stat-card" data-accent="green">
      <div class="stat-label">Total Revenue</div>
      <div class="stat-value stat-accent-green">${{ stats.total_revenue }}</div>
      <div class="stat-sub">{{ stats.purchase_count }} transactions</div>
    </div>
    <div class="stat-card" data-accent="purple">
      <div class="stat-label">Total Referrals</div>
      <div class="stat-value">{{ stats.total_referrals }}</div>
      <div class="stat-sub">All time users</div>
    </div>
    <div class="stat-card" data-accent="purple">
      <div class="stat-label">New Referrals</div>
      <div class="stat-value stat-accent-purple">{{ stats.monthly_referrals }}</div>
      <div class="stat-sub">This month</div>
    </div>
    <div class="stat-card" data-accent="green">
      <div class="stat-label">Monthly Revenue</div>
      <div class="stat-value stat-accent-green">${{ stats.monthly_revenue }}</div>
      <div class="stat-sub">This month</div>
    </div>
    <div class="stat-card" data-accent="yellow">
      <div class="stat-label">Avg Earning / Ref</div>
      <div class="stat-value stat-accent-yellow">${{ stats.avg_earning_per_ref }}</div>
      <div class="stat-sub">Per referrer</div>
    </div>
  </div>

  <!-- ── Flowers of week/month ── -->
  <div class="section-title">🌸 Top Flowers</div>
  <div class="flower-grid">
    <div class="flower-card">
      <div class="flower-emoji">🏆</div>
      <div class="flower-info">
        <div class="flower-period">Flower of the Week</div>
        {% if stats.flower_week %}
        <div class="flower-name">{{ stats.flower_week.display_name }}</div>
        <div class="flower-count">{{ stats.flower_week.invite_count }} referrals this week</div>
        {% else %}
        <div class="flower-name" style="color:var(--muted)">No data yet</div>
        {% endif %}
      </div>
    </div>
    <div class="flower-card">
      <div class="flower-emoji">🌟</div>
      <div class="flower-info">
        <div class="flower-period">Flower of the Month</div>
        {% if stats.flower_month %}
        <div class="flower-name">{{ stats.flower_month.display_name }}</div>
        <div class="flower-count">{{ stats.flower_month.invite_count }} referrals this month</div>
        {% else %}
        <div class="flower-name" style="color:var(--muted)">No data yet</div>
        {% endif %}
      </div>
    </div>
  </div>

  <!-- ── Charts ── -->
  <div class="section-title">Trends</div>
  <div class="charts-grid">
    <div class="chart-card">
      <div class="chart-title">Monthly Referrals</div>
      <canvas id="refChart" height="200"></canvas>
    </div>
    <div class="chart-card">
      <div class="chart-title">Monthly Revenue ($)</div>
      <canvas id="revChart" height="200"></canvas>
    </div>
  </div>

  <!-- ── Tables ── -->
  <div class="section-title">Leaderboards & Activity</div>
  <div class="tables-grid">

    <!-- Top Referrers -->
    <div class="table-card">
      <div class="table-head">
        🥇 Who Brought Most Clients
      </div>
      <table>
        <thead><tr><th>#</th><th>User</th><th>Invites</th><th>Earned</th></tr></thead>
        <tbody>
          {% for r in stats.top_referrers %}
          <tr>
            <td><span class="rank-num {% if loop.index <= 3 %}top{% endif %}">{{ loop.index }}</span></td>
            <td>{{ r.display_name }}</td>
            <td><span class="pill pill-accent">{{ r.invite_count }}</span></td>
            <td class="stat-accent-green">${{ "%.2f"|format(r.total_earned) }}</td>
          </tr>
          {% else %}
          <tr class="empty-row"><td colspan="4">No referrals yet</td></tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

    <!-- Monthly Payouts -->
    <div class="table-card">
      <div class="table-head">💸 Monthly Payouts by Referrer</div>
      <table>
        <thead><tr><th>User</th><th>Sales</th><th>Total Paid</th></tr></thead>
        <tbody>
          {% for p in stats.monthly_payouts %}
          <tr>
            <td>{{ p.display_name }}</td>
            <td><span class="pill pill-muted">{{ p.sale_count }}</span></td>
            <td class="stat-accent-green">${{ "%.2f"|format(p.monthly_total) }}</td>
          </tr>
          {% else %}
          <tr class="empty-row"><td colspan="3">No sales this month</td></tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

    <!-- Most Active Content Posters -->
    <div class="table-card">
      <div class="table-head">📣 Most Active in Publishing</div>
      <table>
        <thead><tr><th>User</th><th>Posts / Shares</th></tr></thead>
        <tbody>
          {% for p in stats.top_posters %}
          <tr>
            <td>{{ p.display_name }}</td>
            <td><span class="pill pill-yellow">{{ p.post_count }}</span></td>
          </tr>
          {% else %}
          <tr class="empty-row"><td colspan="2">No activity tracked yet</td></tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

    <!-- Recent Sales -->
    <div class="table-card">
      <div class="table-head">🛒 Recent Sales</div>
      <table>
        <thead><tr><th>Product</th><th>Amt</th><th>Referrer</th></tr></thead>
        <tbody>
          {% for s in stats.recent_sales %}
          <tr>
            <td>{{ s.product_name }}</td>
            <td class="stat-accent-green">${{ "%.2f"|format(s.amount) }}</td>
            <td>
              {% if s.referrer_id %}
              <span class="pill pill-accent">{{ s.referrer_name }}</span>
              {% else %}
              <span class="pill pill-muted">Organic</span>
              {% endif %}
            </td>
          </tr>
          {% else %}
          <tr class="empty-row"><td colspan="3">No sales yet</td></tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

  </div>
</main>

<footer>
  Auto-refreshes every 30s &nbsp;·&nbsp; ReferralBot Admin &nbsp;·&nbsp;
  <span style="color: var(--accent3)">●</span> System Online
</footer>

<script>
const monthly = {{ stats.monthly_data | tojson }};
const labels  = monthly.map(d => d.month);
const refs    = monthly.map(d => d.referrals);
const revs    = monthly.map(d => d.revenue);

const gridColor = 'rgba(255,255,255,0.04)';
const fontColor = '#6e7191';
const baseOpts  = {
  responsive: true,
  plugins: {
    legend: { display: false },
    tooltip: { backgroundColor: '#191c27', titleColor: '#e8eaf2', bodyColor: '#6e7191', borderColor: '#252837', borderWidth: 1 }
  },
  scales: {
    x: { grid: { color: gridColor }, ticks: { color: fontColor, font: { family: 'DM Mono', size: 11 } } },
    y: { grid: { color: gridColor }, ticks: { color: fontColor, font: { family: 'DM Mono', size: 11 } }, beginAtZero: true }
  }
};

new Chart(document.getElementById('refChart'), {
  type: 'bar',
  data: {
    labels,
    datasets: [{
      data: refs,
      backgroundColor: 'rgba(124,106,247,0.6)',
      borderColor: 'rgba(124,106,247,1)',
      borderWidth: 1.5,
      borderRadius: 6,
    }]
  },
  options: baseOpts
});

new Chart(document.getElementById('revChart'), {
  type: 'line',
  data: {
    labels,
    datasets: [{
      data: revs,
      borderColor: '#43e8a4',
      backgroundColor: 'rgba(67,232,164,0.08)',
      borderWidth: 2,
      pointBackgroundColor: '#43e8a4',
      pointRadius: 4,
      fill: true,
      tension: 0.4
    }]
  },
  options: baseOpts
});

setTimeout(() => location.reload(), 30000);
</script>
</body>
</html>
"""


@app.route('/')
def dashboard():
    stats = get_dashboard_stats()
    return render_template_string(DASHBOARD_HTML, stats=stats)


@app.route('/api/stats')
def api_stats():
    return jsonify(get_dashboard_stats())


@app.route('/export')
def export_data():
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Referrals"
        cursor.execute("SELECT * FROM referrals")
        ws1.append([d[0] for d in cursor.description])
        for row in cursor.fetchall(): ws1.append(list(row))

        ws2 = wb.create_sheet("Purchases")
        cursor.execute("SELECT * FROM purchases")
        ws2.append([d[0] for d in cursor.description])
        for row in cursor.fetchall(): ws2.append(list(row))

        ws3 = wb.create_sheet("Link Clicks")
        cursor.execute("SELECT * FROM link_clicks")
        ws3.append([d[0] for d in cursor.description])
        for row in cursor.fetchall(): ws3.append(list(row))

        ws4 = wb.create_sheet("Content Posts")
        cursor.execute("SELECT * FROM content_posts")
        ws4.append([d[0] for d in cursor.description])
        for row in cursor.fetchall(): ws4.append(list(row))

        conn.close()
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'referral_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        return f"Error: {e}", 500


def run_flask():
    serve(app, host='0.0.0.0', port=5000)


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

if __name__ == '__main__':
    if not BOT_TOKEN:
        print("CRITICAL: BOT_TOKEN missing in .env")
        exit(1)

    init_db()
    print("✅ Database initialized.")

    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()
    print("✅ Dashboard running on http://0.0.0.0:5000 via Waitress")

    print("✅ Starting Telegram bot polling…")
    t_request = HTTPXRequest(connect_timeout=60.0, read_timeout=60.0)
    application = Application.builder().token(BOT_TOKEN).request(t_request).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CallbackQueryHandler(button_handler))
    application.add_handler(CommandHandler("approve", approve_sale))
    application.add_handler(CommandHandler("logpost", log_post_cmd))
    application.add_error_handler(error_handler)

    try:
        application.run_polling(allowed_updates=Update.ALL_TYPES)
    except Exception as e:
        print(f"CRITICAL: {e}")